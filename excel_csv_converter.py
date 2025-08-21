#!/usr/bin/env python3
"""
Excel ↔ CSV Converter
Lightweight cross-platform converter for Excel (.xls, .xlsx) and CSV files
Handles large files efficiently with minimal memory usage
"""

import os
import sys
import csv
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError
import xlrd
import warnings

# Suppress warnings from xlrd
warnings.filterwarnings("ignore", category=UserWarning, module="xlrd")

class ExcelCSVConverter:
    def __init__(self):
        self.supported_excel = ('.xlsx', '.xlsm', '.xls')
        self.supported_csv = ('.csv', '.txt')
        
    def excel_to_csv(self, excel_path, csv_path=None, sheet_index=0):
        """
        Convert Excel file to CSV with UTF-8 encoding.
        Handles large files efficiently using iterators.
        """
        excel_path = Path(excel_path)
        
        if csv_path is None:
            csv_path = excel_path.with_suffix('.csv')
        
        try:
            file_ext = excel_path.suffix.lower()
            
            if file_ext == '.xls':
                # Handle older .xls files
                workbook = xlrd.open_workbook(str(excel_path), encoding_override='utf-8')
                sheet = workbook.sheet_by_index(sheet_index)
                
                with open(csv_path, 'w', newline='', encoding='utf-8-sig', errors='replace') as csvfile:
                    writer = csv.writer(csvfile)
                    for row_idx in range(sheet.nrows):
                        row = []
                        for col_idx in range(sheet.ncols):
                            cell = sheet.cell(row_idx, col_idx)
                            # Handle different cell types
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                # Convert Excel date to readable format
                                from xlrd import xldate_as_datetime
                                try:
                                    dt = xldate_as_datetime(cell.value, workbook.datemode)
                                    row.append(dt.strftime('%Y-%m-%d %H:%M:%S'))
                                except:
                                    row.append(str(cell.value))
                            else:
                                # Ensure proper string handling for all characters
                                cell_value = str(cell.value)
                                # Remove any null bytes that might cause issues
                                cell_value = cell_value.replace('\x00', '')
                                row.append(cell_value)
                        writer.writerow(row)
            else:
                # Handle .xlsx files with openpyxl (memory efficient for large files)
                workbook = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
                sheet = workbook.worksheets[sheet_index]
                
                with open(csv_path, 'w', newline='', encoding='utf-8-sig', errors='replace') as csvfile:
                    writer = csv.writer(csvfile)
                    # Use iter_rows for memory efficiency with large files
                    for row in sheet.iter_rows(values_only=True):
                        # Convert None values to empty strings and handle all unicode properly
                        row_data = []
                        for cell in row:
                            if cell is None:
                                row_data.append('')
                            else:
                                # Ensure proper string conversion for all types
                                cell_str = str(cell)
                                # Remove any null bytes
                                cell_str = cell_str.replace('\x00', '')
                                row_data.append(cell_str)
                        writer.writerow(row_data)
                
                workbook.close()
            
            return True, f"Successfully converted to: {csv_path}"
            
        except Exception as e:
            return False, f"Error converting Excel to CSV: {str(e)}"
    
    def csv_to_excel(self, csv_path, excel_path=None, xlsx=True):
        """
        Convert CSV file to Excel format.
        Handles complex CSVs with JSON data, Unicode characters, and special formatting.
        Uses pandas as primary method for better handling of complex data.
        """
        csv_path = Path(csv_path)
        
        if excel_path is None:
            extension = '.xlsx' if xlsx else '.xls'
            excel_path = csv_path.with_suffix(extension)
        
        # Try pandas first if available (usually more robust)
        try:
            import pandas as pd
            print("Using pandas for conversion (recommended for complex files)...")
            
            # Read CSV with pandas - very robust for complex data
            df = pd.read_csv(
                csv_path, 
                encoding='utf-8',
                encoding_errors='replace',
                on_bad_lines='warn',  # Warn but don't fail on bad lines
                engine='python',  # Python engine handles complex quotes better
                sep=None,  # Auto-detect separator
                dtype=str,  # Keep everything as strings to avoid conversion issues
                keep_default_na=False,  # Don't convert strings to NaN
                na_values=[''],  # Only treat empty strings as NaN
                quoting=csv.QUOTE_MINIMAL
            )
            
            # Handle large cells that exceed Excel's limit
            for col in df.columns:
                df[col] = df[col].apply(lambda x: x[:32767] if isinstance(x, str) and len(x) > 32767 else x)
            
            # Write to Excel
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Sheet1', index=False)
            
            row_count = len(df)
            print(f"Successfully converted {row_count} rows using pandas")
            return True, f"Successfully converted to: {excel_path}\n(Processed {row_count} rows using optimized method)"
            
        except ImportError:
            print("Pandas not available, using standard method...")
        except Exception as pandas_error:
            print(f"Pandas conversion failed: {pandas_error}, trying standard method...")
        
        # Fallback to original method if pandas fails or isn't available
        try:
            # Create a new workbook
            workbook = Workbook(write_only=True)  # write_only mode for memory efficiency
            sheet = workbook.create_sheet('Sheet1')
            
            # For complex CSVs, always try UTF-8 first
            encoding_used = 'utf-8'
            
            # Read the entire file to handle complex parsing
            print(f"Reading CSV file: {csv_path}")
            with open(csv_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
            
            # Use csv.reader with specific settings for complex data
            import io
            csvfile = io.StringIO(content)
            
            # Configure CSV reader for complex data
            # Set a larger field size limit for JSON data
            import sys
            maxInt = sys.maxsize
            while True:
                try:
                    csv.field_size_limit(maxInt)
                    break
                except OverflowError:
                    maxInt = int(maxInt/10)  # Reduce by factor of 10 until it works
            
            # Use excel dialect which handles quotes better
            reader = csv.reader(csvfile, dialect='excel', 
                              quoting=csv.QUOTE_MINIMAL,
                              escapechar='\\',
                              doublequote=True)
            
            row_num = 0
            error_count = 0
            max_errors = 50  # Increased tolerance for errors
            
            for row in reader:
                row_num += 1
                
                if row_num % 100 == 0:
                    print(f"Processing row {row_num}...")
                
                # Skip completely empty rows
                if not row or all(cell == '' for cell in row):
                    continue
                
                processed_row = []
                
                for cell_idx, cell in enumerate(row):
                    try:
                        if cell == '' or cell is None:
                            processed_row.append(None)
                        else:
                            # Clean the cell value
                            cell_str = str(cell)
                            
                            # Remove null bytes and other problematic characters
                            cell_str = cell_str.replace('\x00', '')
                            
                            # Remove other control characters except tab, newline, carriage return
                            cleaned = ''.join(char for char in cell_str 
                                            if ord(char) >= 32 or char in '\t\n\r')
                            
                            # Truncate extremely long cells (Excel has a 32,767 character limit)
                            if len(cleaned) > 32767:
                                cleaned = cleaned[:32764] + '...'
                                print(f"Warning: Truncated long cell at row {row_num}, column {cell_idx + 1}")
                            
                            processed_row.append(cleaned)
                    except Exception as e:
                        print(f"Error processing cell at row {row_num}, column {cell_idx + 1}: {e}")
                        processed_row.append('')  # Add empty string on error
                
                try:
                    sheet.append(processed_row)
                except Exception as e:
                    error_count += 1
                    print(f"Error at row {row_num}: {str(e)[:100]}")
                    
                    # Try again with all strings
                    try:
                        safe_row = []
                        for cell in processed_row:
                            if cell is None:
                                safe_row.append('')
                            else:
                                # Extra safety - remove ALL control characters
                                safe_cell = ''.join(char for char in str(cell) 
                                                  if ord(char) >= 32 or char in '\t\n\r')
                                safe_row.append(safe_cell[:32767])  # Ensure within Excel limits
                        sheet.append(safe_row)
                        print(f"Row {row_num} saved with extra cleaning")
                    except Exception as e2:
                        print(f"Row {row_num} failed completely: {str(e2)[:100]}")
                        if error_count > max_errors:
                            # Try to save what we have so far
                            print(f"Too many errors ({error_count}). Saving partial file...")
                            break
                        # Skip this row and continue
                        continue
            
            print(f"Processed {row_num} rows total with {error_count} errors")
            
            # Save the workbook
            print("Saving Excel file...")
            workbook.save(str(excel_path))
            workbook.close()
            
            if error_count > 0:
                return True, f"Successfully converted to: {excel_path}\n(Processed {row_num} rows with {error_count} errors - some data may have been cleaned or skipped)"
            else:
                return True, f"Successfully converted to: {excel_path}\n(Processed {row_num} rows successfully)"
            
        except Exception as e:
            error_msg = f"Error converting CSV to Excel: {str(e)}"
            print(f"Full error: {error_msg}")
            return False, error_msg
    
    def get_excel_info(self, excel_path):
        """Get information about Excel file (sheet names, dimensions, etc.)"""
        try:
            excel_path = Path(excel_path)
            file_ext = excel_path.suffix.lower()
            
            if file_ext == '.xls':
                workbook = xlrd.open_workbook(str(excel_path))
                sheets = []
                for i in range(workbook.nsheets):
                    sheet = workbook.sheet_by_index(i)
                    sheets.append({
                        'name': sheet.name,
                        'rows': sheet.nrows,
                        'cols': sheet.ncols
                    })
                return sheets
            else:
                workbook = openpyxl.load_workbook(str(excel_path), read_only=True)
                sheets = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheets.append({
                        'name': sheet_name,
                        'rows': sheet.max_row,
                        'cols': sheet.max_column
                    })
                workbook.close()
                return sheets
        except Exception as e:
            return None

class ConverterGUI:
    def __init__(self):
        self.converter = ExcelCSVConverter()
        self.root = tk.Tk()
        self.root.title("Excel ↔ CSV Converter")
        self.root.geometry("500x300")
        self.root.resizable(False, False)
        
        # Set icon for different platforms
        if sys.platform == "win32":
            self.root.iconbitmap(default='')  # Default Windows icon
        
        self.setup_ui()
        
    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Title
        title = ttk.Label(main_frame, text="Excel ↔ CSV Converter", font=('Helvetica', 16, 'bold'))
        title.grid(row=0, column=0, columnspan=2, pady=10)
        
        # Instructions
        instructions = ttk.Label(main_frame, text="Select a file to convert:", font=('Helvetica', 10))
        instructions.grid(row=1, column=0, columnspan=2, pady=5)
        
        # File info label
        self.file_info = ttk.Label(main_frame, text="No file selected", font=('Helvetica', 9))
        self.file_info.grid(row=2, column=0, columnspan=2, pady=10)
        
        # Sheet selection frame (hidden initially)
        self.sheet_frame = ttk.Frame(main_frame)
        self.sheet_frame.grid(row=3, column=0, columnspan=2, pady=5)
        self.sheet_frame.grid_remove()
        
        ttk.Label(self.sheet_frame, text="Select sheet:").pack(side=tk.LEFT, padx=5)
        self.sheet_var = tk.StringVar()
        self.sheet_combo = ttk.Combobox(self.sheet_frame, textvariable=self.sheet_var, width=30, state="readonly")
        self.sheet_combo.pack(side=tk.LEFT)
        
        # Buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=4, column=0, columnspan=2, pady=20)
        
        # Convert Excel to CSV button
        excel_to_csv_btn = ttk.Button(
            button_frame,
            text="Convert Excel to CSV",
            command=self.convert_excel_to_csv,
            width=20
        )
        excel_to_csv_btn.grid(row=0, column=0, padx=5, pady=5)
        
        # Convert CSV to Excel button
        csv_to_excel_btn = ttk.Button(
            button_frame,
            text="Convert CSV to Excel",
            command=self.convert_csv_to_excel,
            width=20
        )
        csv_to_excel_btn.grid(row=0, column=1, padx=5, pady=5)
        
        # Batch convert button
        batch_btn = ttk.Button(
            button_frame,
            text="Batch Convert Folder",
            command=self.batch_convert,
            width=20
        )
        batch_btn.grid(row=1, column=0, columnspan=2, pady=5)
        
        # Status bar
        self.status_bar = ttk.Label(main_frame, text="Ready", relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        
    def convert_excel_to_csv(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel file",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm *.xls"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.file_info.config(text=f"Selected: {Path(file_path).name}")
            
            # Get sheet information
            sheets = self.converter.get_excel_info(file_path)
            if sheets and len(sheets) > 1:
                # Show sheet selection
                self.sheet_frame.grid()
                sheet_names = [f"{s['name']} ({s['rows']} rows, {s['cols']} cols)" for s in sheets]
                self.sheet_combo['values'] = sheet_names
                self.sheet_combo.current(0)
                
                # Wait for user to select sheet
                result = messagebox.askquestion("Multiple Sheets", 
                                               "This file has multiple sheets. Convert the first sheet?")
                if result == 'no':
                    return
                sheet_index = 0
            else:
                sheet_index = 0
                self.sheet_frame.grid_remove()
            
            # Ask for output location
            output_path = filedialog.asksaveasfilename(
                title="Save CSV as",
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialfile=Path(file_path).stem + ".csv"
            )
            
            if output_path:
                self.status_bar.config(text="Converting...")
                self.root.update()
                
                success, message = self.converter.excel_to_csv(file_path, output_path, sheet_index)
                
                if success:
                    self.status_bar.config(text="Conversion successful!")
                    messagebox.showinfo("Success", message)
                else:
                    self.status_bar.config(text="Conversion failed")
                    messagebox.showerror("Error", message)
    
    def convert_csv_to_excel(self):
        file_path = filedialog.askopenfilename(
            title="Select CSV file",
            filetypes=[
                ("CSV files", "*.csv"),
                ("Text files", "*.txt"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.file_info.config(text=f"Selected: {Path(file_path).name}")
            self.sheet_frame.grid_remove()
            
            # Ask for format
            format_choice = messagebox.askyesno("Excel Format", 
                                               "Save as .xlsx? (Yes for .xlsx, No for .xls)")
            
            # Ask for output location
            ext = ".xlsx" if format_choice else ".xls"
            output_path = filedialog.asksaveasfilename(
                title="Save Excel as",
                defaultextension=ext,
                filetypes=[("Excel files", f"*{ext}"), ("All files", "*.*")],
                initialfile=Path(file_path).stem + ext
            )
            
            if output_path:
                self.status_bar.config(text="Converting...")
                self.root.update()
                
                success, message = self.converter.csv_to_excel(file_path, output_path, format_choice)
                
                if success:
                    self.status_bar.config(text="Conversion successful!")
                    messagebox.showinfo("Success", message)
                else:
                    self.status_bar.config(text="Conversion failed")
                    messagebox.showerror("Error", message)
    
    def batch_convert(self):
        folder_path = filedialog.askdirectory(title="Select folder for batch conversion")
        
        if folder_path:
            # Ask what type of conversion
            choice = messagebox.askyesno("Batch Conversion", 
                                        "Convert Excel to CSV? (No = CSV to Excel)")
            
            folder = Path(folder_path)
            converted = 0
            failed = 0
            
            if choice:  # Excel to CSV
                files = list(folder.glob("*.xlsx")) + list(folder.glob("*.xls")) + list(folder.glob("*.xlsm"))
                for file_path in files:
                    output_path = file_path.with_suffix('.csv')
                    success, _ = self.converter.excel_to_csv(file_path, output_path)
                    if success:
                        converted += 1
                    else:
                        failed += 1
            else:  # CSV to Excel
                files = list(folder.glob("*.csv"))
                for file_path in files:
                    output_path = file_path.with_suffix('.xlsx')
                    success, _ = self.converter.csv_to_excel(file_path, output_path)
                    if success:
                        converted += 1
                    else:
                        failed += 1
            
            message = f"Batch conversion complete!\nConverted: {converted} files\nFailed: {failed} files"
            messagebox.showinfo("Batch Conversion Results", message)
            self.status_bar.config(text=f"Batch conversion complete: {converted} succeeded, {failed} failed")
    
    def run(self):
        self.root.mainloop()

def main():
    # Check if running with command line arguments
    if len(sys.argv) > 1:
        converter = ExcelCSVConverter()
        input_file = Path(sys.argv[1])
        
        if not input_file.exists():
            print(f"Error: File '{input_file}' not found")
            sys.exit(1)
        
        # Determine conversion type based on file extension
        if input_file.suffix.lower() in converter.supported_excel:
            # Convert Excel to CSV
            output_file = input_file.with_suffix('.csv') if len(sys.argv) < 3 else Path(sys.argv[2])
            success, message = converter.excel_to_csv(input_file, output_file)
        elif input_file.suffix.lower() in converter.supported_csv:
            # Convert CSV to Excel
            output_file = input_file.with_suffix('.xlsx') if len(sys.argv) < 3 else Path(sys.argv[2])
            success, message = converter.csv_to_excel(input_file, output_file)
        else:
            print(f"Error: Unsupported file type '{input_file.suffix}'")
            sys.exit(1)
        
        print(message)
        sys.exit(0 if success else 1)
    else:
        # Run GUI
        app = ConverterGUI()
        app.run()

if __name__ == "__main__":
    main()